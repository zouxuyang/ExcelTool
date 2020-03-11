from tkinter import *
from tkinter import messagebox
import tkinter.filedialog
import openpyxl
import time
import os
import threading
import datetime


class App(Frame):
    def __init__(self, master):
        frame = Frame(master, width=120, height=265, bg='white')
        frame.grid(row=0, column=0)
        frame1 = Frame(master, width=235, height=265, bg='red')
        frame1.grid(row=0, column=1)
        self.button1 = Button(frame, text='模板文件确认',
                              command=self.Btn1, bg="lightcyan")
        self.button1.place(relx=0.1, rely=0.05, relheight=0.1, relwidth=0.8)
        self.button2 = Button(frame, text='数据文件夹',
                              command=self.Btn2, bg="lightcyan")
        self.button2.place(relx=0.1, rely=0.2, relheight=0.1, relwidth=0.8)
        self.label = Label(frame, text='统计总天数', bg="lightcyan")
        self.label.place(relx=0.1, rely=0.35, relheight=0.05, relwidth=0.8)
        self.entry = Entry(frame, bg="lightcyan")  # 创建文本框
        self.entry.place(relx=0.1, rely=0.40, relheight=0.1, relwidth=0.8)

        self.button3 = Button(frame, text='开始合并',
                              command=self.Btn3, bg="lightcyan")
        self.button3.place(relx=0.1, rely=0.55, relheight=0.1, relwidth=0.8)
        self.txt = Text(frame1, bg="azure")
        self.txt.place(width=235, height=270)

    def progress_bar(self):
        days = 365
        for i in range(days):
            s = "\n合并进行中：{0}%".format(round((i+1)*100/days))+"\n"
            self.txt.delete(7.0, "end")
            self.txt.insert(END, s)
            self.txt.see(END)
            time.sleep(0.003)
        self.txt.insert(END, "文件合并已完成\n")
        self.txt.see(END)
        os.system('explorer.exe /n, D:\\targetFiles\\')

    def fileCreate(self):
        Cfile = "D:\\targetFiles\\"
        Cfile = Cfile.strip()
        Cfile = Cfile.rstrip("\\")
        isExists = os.path.exists(Cfile)
        if not isExists:
            os.makedirs(Cfile)
            folder_path = '目录创建成功\n'+'文件存储路径：'+Cfile+"\n"
            self.txt.insert(END, folder_path)
            self.txt.see(END)
            return True
        else:
            folder_path = '目录已存在，文件存储路径：'+Cfile+"\n"
            self.txt.insert(END, folder_path)
            self.txt.see(END)
            return False

    def Btn1(self):
        self.filename1 = tkinter.filedialog.askopenfilename()
        if self.filename1 != '':
            s = '您选择的模板文件是'+self.filename1+"\n"
            self.txt.insert(END, s)
            self.txt.see(END)
            print(self.filename1)
        else:
            s = '您没有选择任何文件'+"\n"
            self.txt.insert(END, s)
            self.txt.see(END)

    def Btn2(self):
        self.Folderpath = tkinter.filedialog.askdirectory()
        if self.Folderpath != '':
            s = '您选择的文件夹是'+self.Folderpath+"\n"
            self.txt.insert(END, s)
            self.txt.see(END)
        else:
            s = '您没有选择任何文件夹'+"\n"
            self.txt.insert(END, s)
            self.txt.see(END)

    def Btn3_Func(self):
        self.fileCreate()
        wb1 = openpyxl.load_workbook(self.filename1)
        ws1 = wb1.active
        SumDate = self.entry.get()
        filenames = os.listdir(self.Folderpath)
        for filename in filenames:
            wb = openpyxl.load_workbook((self.Folderpath+"/"+filename))
            ws = wb.active
            for i in range(0, int(SumDate)):
                ws1.append([ws['B2'].value, ws['B1'].value, ws['B3'].value, ws['A'+str(6+2*i)].value, ws['C'+str(6+2*i)].value, ws['D'+str(6+2*i)].value,
                            ws['E'+str(6+2*i)].value, ws['F'+str(6+2*i)].value, ws['G'+str(6+2*i)].value, ws['H'+str(6+2*i)].value, ws['I'+str(6+2*i)].value, ws['J'+str(6+2*i)].value])
        Cfile = "D:\\targetFiles\\"
        menu = '%s%s%s' % (
            Cfile, datetime.datetime.now().strftime('%Y%m%d%H%M%S'), '.xlsx')
        wb1.save(menu)

    def Btn3(self):
        th1 = threading.Thread(target=self.progress_bar)
        th1.setDaemon(True)
        th1.start()
        th2 = threading.Thread(target=self.Btn3_Func)
        th2.setDaemon(True)
        th2.start()


if __name__ == '__main__':
    root = Tk()
    root.title('神隐办公自动化')
    root.geometry('360x270')
    App(root)
    root.mainloop()
